import argparse
import csv
import json
import os
import sys
import tkinter as tk
from dataclasses import dataclass
from datetime import datetime
from tkinter import filedialog, messagebox, ttk
from tkinter import font as tkfont
from typing import Any, Iterable


APP_TITLE = "Lakshmipur Government College – Teacher Phone Book"
SETTINGS_FILE = "phonebook_settings.json"
DEFAULT_DATA_CANDIDATES = (
    "teachers.csv",
    "teacher_phonebook.csv",
    "teachers_phonebook.csv",
    "teacher_phone_book.csv",
    "teachers.xlsx",
    "teacher_phonebook.xlsx",
    "phonebook.xlsx",
    "PhoneBook.xlsx",
)


def set_windows_dpi_awareness() -> None:
    if sys.platform != "win32":
        return
    try:
        import ctypes

        try:
            ctypes.windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            ctypes.windll.user32.SetProcessDPIAware()
    except Exception:
        return


def load_settings() -> dict[str, Any]:
    try:
        with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except FileNotFoundError:
        return {}
    except Exception:
        return {}


def save_settings(data: dict[str, Any]) -> None:
    try:
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _make_unique_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for header in headers:
        base = header.strip() or "Column"
        if base not in seen:
            seen[base] = 1
            out.append(base)
            continue
        seen[base] += 1
        out.append(f"{base} ({seen[base]})")
    return out


def _is_row_empty(row: dict[str, str]) -> bool:
    return all(not (value or "").strip() for value in row.values())


def read_csv_records(path: str) -> tuple[list[str], list[dict[str, str]]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=[",", ";", "\t", "|"])
        except Exception:
            dialect = csv.excel

        reader = csv.DictReader(f, dialect=dialect)
        if not reader.fieldnames:
            raise ValueError("CSV file has no header row.")

        raw_headers = [h.strip() for h in reader.fieldnames]
        headers = _make_unique_headers(raw_headers)

        records: list[dict[str, str]] = []
        for row in reader:
            normalized: dict[str, str] = {}
            for original, header in zip(raw_headers, headers):
                value = row.get(original, "")
                normalized[header] = "" if value is None else str(value).strip()
            if _is_row_empty(normalized):
                continue
            records.append(normalized)

    return headers, records


def read_xlsx_records(path: str) -> tuple[list[str], list[dict[str, str]]]:
    try:
        import openpyxl  # type: ignore
    except Exception as e:
        raise RuntimeError("Reading .xlsx requires the 'openpyxl' package.") from e

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    first_cell = ws.cell(row=1, column=1).value
    looks_like_csv_in_one_column = (
        ws.max_column == 1
        and isinstance(first_cell, str)
        and any(sep in first_cell for sep in (",", ";", "\t", "|"))
    )

    if looks_like_csv_in_one_column:
        lines: list[str] = []
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=1).value
            if value is None:
                continue
            text = str(value).strip()
            if text:
                lines.append(text)
        if not lines:
            raise ValueError("Excel file is empty.")

        sample = "\n".join(lines[:30])
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=[",", ";", "\t", "|"])
        except Exception:
            dialect = csv.excel

        reader = csv.DictReader(lines, dialect=dialect)
        if not reader.fieldnames:
            raise ValueError("Excel file has no header row.")

        raw_headers = [h.strip() for h in reader.fieldnames]
        headers = _make_unique_headers(raw_headers)

        records: list[dict[str, str]] = []
        for row in reader:
            normalized: dict[str, str] = {}
            for original, header in zip(raw_headers, headers):
                value = row.get(original, "")
                normalized[header] = "" if value is None else str(value).strip()
            if _is_row_empty(normalized):
                continue
            records.append(normalized)
        return headers, records

    headers_row = None
    for r in range(1, min(ws.max_row, 20) + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if any(v is not None and str(v).strip() for v in values):
            headers_row = r
            break

    if headers_row is None:
        raise ValueError("Excel file is empty.")

    raw_headers = [
        "" if v is None else str(v).strip() for v in [ws.cell(row=headers_row, column=c).value for c in range(1, ws.max_column + 1)]
    ]
    if not any(raw_headers):
        raise ValueError("Excel header row is empty.")

    headers = _make_unique_headers(raw_headers)

    records: list[dict[str, str]] = []
    for r in range(headers_row + 1, ws.max_row + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        record: dict[str, str] = {}
        for idx, header in enumerate(headers):
            value = values[idx] if idx < len(values) else None
            record[header] = "" if value is None else str(value).strip()
        if _is_row_empty(record):
            continue
        records.append(record)

    return headers, records


def read_tabular_records(path: str) -> tuple[list[str], list[dict[str, str]]]:
    lower = path.lower()
    if lower.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        return read_xlsx_records(path)
    return read_csv_records(path)


def _is_phone_like(value: str) -> bool:
    text = (value or "").strip()
    if not text:
        return False
    if "@" in text:
        return False
    digits = "".join(ch for ch in text if ch.isdigit())
    return 7 <= len(digits) <= 15


def guess_phone_column(columns: Iterable[str], records: Iterable[dict[str, str]] | None = None) -> str | None:
    columns_list = list(columns)

    keyword_hits: list[str] = []
    for col in columns_list:
        key = col.strip().casefold()
        if any(
            token in key
            for token in (
                "phone",
                "mobile",
                "tel",
                "telephone",
                "contact",
                "cell",
                "number",
                "no.",
                "ফোন",
                "মোবাইল",
                "যোগাযোগ",
                "নম্বর",
                "নং",
            )
        ):
            keyword_hits.append(col)
    if keyword_hits:
        return keyword_hits[0]

    if not records:
        return None

    best_col: str | None = None
    best_score = 0.0
    max_rows = 250

    for col in columns_list:
        checked = 0
        phone_like = 0
        for idx, record in enumerate(records):
            if idx >= max_rows:
                break
            value = record.get(col, "").strip()
            if not value:
                continue
            checked += 1
            if _is_phone_like(value):
                phone_like += 1
        if checked == 0:
            continue
        score = phone_like / checked
        if score > best_score:
            best_score = score
            best_col = col

    return best_col if best_score >= 0.35 else None


def detect_initial_csv_path(cli_path: str | None) -> str | None:
    if cli_path and os.path.isfile(cli_path):
        return cli_path

    settings = load_settings()
    last_path = settings.get("last_csv")
    if isinstance(last_path, str) and os.path.isfile(last_path):
        return last_path

    for name in DEFAULT_DATA_CANDIDATES:
        if os.path.isfile(name):
            return name

    try:
        csv_files: list[str] = []
        sample_file: str | None = None
        for name in os.listdir("."):
            if not (name.lower().endswith(".csv") and os.path.isfile(name)):
                continue
            if name.lower() == "teachers_sample.csv":
                sample_file = name
                continue
            csv_files.append(name)

        if csv_files:
            return max(csv_files, key=lambda p: os.path.getmtime(p))
        if sample_file:
            return sample_file
    except Exception:
        pass
    return None


@dataclass(frozen=True)
class SortState:
    column: str
    reverse: bool


class PhoneBookApp(tk.Tk):
    def __init__(self, initial_csv: str | None = None) -> None:
        super().__init__()

        self.title(APP_TITLE)
        self.minsize(860, 560)
        self.geometry("1100x700")

        self.csv_path: str | None = None
        self.columns: list[str] = []
        self.all_records: list[dict[str, str]] = []
        self.filtered_records: list[dict[str, str]] = []
        self.phone_column: str | None = None
        self.sort_state: SortState | None = None

        self._filter_job: str | None = None
        self._context_item: str | None = None
        self._context_column: str | None = None

        self.search_var = tk.StringVar()
        self.search_column_var = tk.StringVar(value="All columns")

        self._configure_theme()
        self._build_ui()
        self._wire_events()

        if initial_csv:
            self.load_csv(initial_csv)
        else:
            self._refresh_status("Ready. Open a CSV file to start.")

    def _configure_theme(self) -> None:
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass

        self.configure(padx=10, pady=10)
        style.configure("Toolbar.TFrame", padding=(6, 6))
        style.configure("Status.TLabel", padding=(6, 2))

    def _build_ui(self) -> None:
        self._build_toolbar()
        self._build_filters()
        self._build_table()
        self._build_details()
        self._build_statusbar()

    def _build_toolbar(self) -> None:
        frame = ttk.Frame(self, style="Toolbar.TFrame")
        frame.pack(fill="x")

        self.open_button = ttk.Button(frame, text="Open CSV…", command=self.open_csv_dialog)
        self.open_button.pack(side="left")

        self.reload_button = ttk.Button(frame, text="Reload", command=self.reload_csv, state="disabled")
        self.reload_button.pack(side="left", padx=(8, 0))

        self.export_button = ttk.Button(frame, text="Export Filtered…", command=self.export_filtered, state="disabled")
        self.export_button.pack(side="left", padx=(8, 0))

        ttk.Separator(frame, orient="vertical").pack(side="left", fill="y", padx=10)

        self.help_button = ttk.Button(frame, text="Help", command=self.show_help)
        self.help_button.pack(side="left")

        self.file_label = ttk.Label(frame, text="No file loaded")
        self.file_label.pack(side="right")

    def _build_filters(self) -> None:
        frame = ttk.Frame(self)
        frame.pack(fill="x", pady=(10, 6))

        ttk.Label(frame, text="Search:").pack(side="left")
        self.search_entry = ttk.Entry(frame, textvariable=self.search_var)
        self.search_entry.pack(side="left", fill="x", expand=True, padx=(6, 8))

        ttk.Label(frame, text="In:").pack(side="left")
        self.search_column_combo = ttk.Combobox(
            frame,
            textvariable=self.search_column_var,
            state="readonly",
            values=["All columns"],
            width=18,
        )
        self.search_column_combo.pack(side="left", padx=(6, 8))

        self.clear_button = ttk.Button(frame, text="Clear", command=self.clear_filters)
        self.clear_button.pack(side="left")

        self.count_label = ttk.Label(frame, text="0 records")
        self.count_label.pack(side="right")

    def _build_table(self) -> None:
        outer = ttk.Frame(self)
        outer.pack(fill="both", expand=True)

        self.tree = ttk.Treeview(outer, show="headings", selectmode="browse")
        vsb = ttk.Scrollbar(outer, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(outer, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        outer.grid_rowconfigure(0, weight=1)
        outer.grid_columnconfigure(0, weight=1)

    def _build_details(self) -> None:
        frame = ttk.LabelFrame(self, text="Selected Teacher")
        frame.pack(fill="x", pady=(10, 0))

        self.details_text = tk.Text(frame, height=6, wrap="word", state="disabled")
        self.details_text.pack(side="left", fill="both", expand=True, padx=(6, 0), pady=6)

        actions = ttk.Frame(frame)
        actions.pack(side="right", fill="y", padx=6, pady=6)

        self.copy_phone_button = ttk.Button(actions, text="Copy Phone", command=self.copy_phone, state="disabled")
        self.copy_phone_button.pack(fill="x", pady=(0, 6))

        self.copy_row_button = ttk.Button(actions, text="Copy Row", command=self.copy_row, state="disabled")
        self.copy_row_button.pack(fill="x", pady=(0, 6))

        self.open_file_location_button = ttk.Button(
            actions, text="Open File Location", command=self.open_file_location, state="disabled"
        )
        self.open_file_location_button.pack(fill="x")

    def _build_statusbar(self) -> None:
        self.status_var = tk.StringVar(value="")
        label = ttk.Label(self, textvariable=self.status_var, style="Status.TLabel", anchor="w")
        label.pack(fill="x", pady=(8, 0))

    def _wire_events(self) -> None:
        self.search_var.trace_add("write", lambda *_: self._schedule_filter())
        self.search_column_var.trace_add("write", lambda *_: self._schedule_filter())

        self.tree.bind("<<TreeviewSelect>>", lambda _e: self._on_select())
        self.tree.bind("<Double-1>", lambda _e: self.copy_phone())
        self.tree.bind("<Button-3>", self._show_context_menu)

    def _schedule_filter(self) -> None:
        if self._filter_job:
            try:
                self.after_cancel(self._filter_job)
            except Exception:
                pass
        self._filter_job = self.after(160, self.apply_filters)

    def _refresh_status(self, text: str) -> None:
        self.status_var.set(text)

    def _set_buttons_enabled(self, enabled: bool) -> None:
        state = "normal" if enabled else "disabled"
        self.reload_button.configure(state=state)
        self.export_button.configure(state=state)
        self.open_file_location_button.configure(state=state if self.csv_path else "disabled")

    def open_csv_dialog(self) -> None:
        path = filedialog.askopenfilename(
            title="Open teacher data file",
            filetypes=[
                ("Data files", "*.csv;*.xlsx;*.xlsm;*.xltx;*.xltm"),
                ("CSV files", "*.csv"),
                ("Excel files", "*.xlsx;*.xlsm;*.xltx;*.xltm"),
                ("All files", "*.*"),
            ],
        )
        if not path:
            return
        self.load_csv(path)

    def load_csv(self, path: str) -> None:
        try:
            columns, records = read_tabular_records(path)
        except Exception as e:
            messagebox.showerror("Could not load file", f"{e}")
            return

        self.csv_path = os.path.abspath(path)
        self.columns = columns
        self.all_records = records
        self.filtered_records = list(records)
        self.phone_column = guess_phone_column(columns, records)
        self.sort_state = None

        self._configure_tree_columns()
        self.apply_filters(force_refresh=True)

        self.file_label.configure(text=os.path.basename(self.csv_path))
        self._set_buttons_enabled(True)
        self._refresh_status(f"Loaded {len(self.all_records)} teachers from {os.path.basename(self.csv_path)}")

        settings = load_settings()
        settings["last_csv"] = self.csv_path
        save_settings(settings)

    def reload_csv(self) -> None:
        if not self.csv_path:
            return
        self.load_csv(self.csv_path)

    def clear_filters(self) -> None:
        self.search_var.set("")
        self.search_column_var.set("All columns")
        self.apply_filters(force_refresh=True)

    def apply_filters(self, force_refresh: bool = False) -> None:
        if not self.all_records:
            if force_refresh:
                self._render_records([])
            return

        query = self.search_var.get().strip().lower()
        column_choice = self.search_column_var.get()

        if not query:
            filtered = list(self.all_records)
        else:
            filtered = []
            if column_choice != "All columns" and column_choice in self.columns:
                for record in self.all_records:
                    if query in (record.get(column_choice, "").lower()):
                        filtered.append(record)
            else:
                for record in self.all_records:
                    haystack = " ".join(record.get(c, "") for c in self.columns).lower()
                    if query in haystack:
                        filtered.append(record)

        if self.sort_state and self.sort_state.column in self.columns:
            filtered = self._sorted_records(filtered, self.sort_state.column, self.sort_state.reverse)

        self.filtered_records = filtered
        self._render_records(filtered)

        self.count_label.configure(text=f"{len(filtered)} / {len(self.all_records)} records")
        self.copy_row_button.configure(state="normal" if self._selected_record() else "disabled")
        self.copy_phone_button.configure(
            state="normal" if (self._selected_record() and self.phone_column) else "disabled"
        )

    def _sorted_records(self, records: list[dict[str, str]], column: str, reverse: bool) -> list[dict[str, str]]:
        def key_fn(record: dict[str, str]) -> tuple[int, str]:
            raw = record.get(column, "")
            normalized = raw.strip()
            digits = "".join(ch for ch in normalized if ch.isdigit())
            if len(digits) >= 6:
                return (0, digits)
            return (1, normalized.casefold())

        return sorted(records, key=key_fn, reverse=reverse)

    def _configure_tree_columns(self) -> None:
        self.tree.configure(columns=self.columns)
        self.tree.delete(*self.tree.get_children())

        for col in self.columns:
            self.tree.heading(col, text=col, command=lambda c=col: self._toggle_sort(c))
            self.tree.column(col, width=150, anchor="w")

        self.search_column_combo.configure(values=["All columns"] + self.columns)
        self.search_column_var.set("All columns")

    def _toggle_sort(self, column: str) -> None:
        if not self.filtered_records:
            return
        if self.sort_state and self.sort_state.column == column:
            new_state = SortState(column=column, reverse=not self.sort_state.reverse)
        else:
            new_state = SortState(column=column, reverse=False)
        self.sort_state = new_state
        self.apply_filters(force_refresh=True)

    def _render_records(self, records: list[dict[str, str]]) -> None:
        self.tree.delete(*self.tree.get_children())

        for record in records:
            values = [record.get(col, "") for col in self.columns]
            self.tree.insert("", "end", values=values)

        self._autosize_columns(records)
        self._on_select()

    def _autosize_columns(self, records: list[dict[str, str]]) -> None:
        if not self.columns:
            return
        fnt = tkfont.nametofont("TkDefaultFont")
        max_rows = 80
        for col in self.columns:
            header_w = fnt.measure(col) + 22
            value_w = header_w
            for record in records[:max_rows]:
                value = record.get(col, "")
                value_w = max(value_w, fnt.measure(value) + 22)
            self.tree.column(col, width=min(max(90, value_w), 420))

    def _selected_record(self) -> dict[str, str] | None:
        sel = self.tree.selection()
        if not sel:
            return None
        item = sel[0]
        values = self.tree.item(item, "values")
        if not values or not self.columns:
            return None
        return {col: str(values[idx]) if idx < len(values) else "" for idx, col in enumerate(self.columns)}

    def _on_select(self) -> None:
        record = self._selected_record()
        self._set_details(record)
        self.copy_row_button.configure(state="normal" if record else "disabled")
        self.copy_phone_button.configure(state="normal" if (record and self.phone_column) else "disabled")

    def _set_details(self, record: dict[str, str] | None) -> None:
        self.details_text.configure(state="normal")
        self.details_text.delete("1.0", "end")
        if not record:
            self.details_text.insert("1.0", "Select a teacher from the table to see details.")
        else:
            lines = []
            for col in self.columns:
                value = record.get(col, "")
                if value:
                    lines.append(f"{col}: {value}")
            self.details_text.insert("1.0", "\n".join(lines) if lines else "(No details)")
        self.details_text.configure(state="disabled")

    def copy_phone(self) -> None:
        record = self._selected_record()
        if not record:
            return
        if not self.phone_column:
            messagebox.showinfo("Phone column not found", "No phone/mobile column was detected in your CSV headers.")
            return

        phone = record.get(self.phone_column, "").strip()
        if not phone:
            messagebox.showinfo("No phone number", "The selected teacher record has no phone number in the phone column.")
            return

        self.clipboard_clear()
        self.clipboard_append(phone)
        self._refresh_status(f"Copied phone: {phone}")

    def copy_row(self) -> None:
        record = self._selected_record()
        if not record:
            return
        text = "\t".join(record.get(c, "") for c in self.columns)
        self.clipboard_clear()
        self.clipboard_append(text)
        self._refresh_status("Copied selected row to clipboard (tab-separated).")

    def export_filtered(self) -> None:
        if not self.filtered_records or not self.columns:
            return
        default_name = f"teachers_export_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
        path = filedialog.asksaveasfilename(
            title="Export filtered teachers",
            defaultextension=".csv",
            initialfile=default_name,
            filetypes=[("CSV files", "*.csv")],
        )
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=self.columns)
                writer.writeheader()
                for record in self.filtered_records:
                    writer.writerow({c: record.get(c, "") for c in self.columns})
        except Exception as e:
            messagebox.showerror("Export failed", f"{e}")
            return
        self._refresh_status(f"Exported {len(self.filtered_records)} records to {os.path.basename(path)}")

    def open_file_location(self) -> None:
        if not self.csv_path:
            return
        folder = os.path.dirname(self.csv_path)
        try:
            if sys.platform == "win32":
                os.startfile(folder)  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                import subprocess

                subprocess.run(["open", folder], check=False)
            else:
                import subprocess

                subprocess.run(["xdg-open", folder], check=False)
        except Exception:
            messagebox.showinfo("Open folder", folder)

    def show_help(self) -> None:
        message = (
            "How to use:\n"
            "1) Click 'Open CSV…' and select your teacher list CSV.\n"
            "2) Use the Search box to find by name/department/phone.\n"
            "3) Click a column header to sort.\n"
            "4) Select a row, then use 'Copy Phone' or 'Copy Row'.\n\n"
            "CSV format tips:\n"
            "- First row should be headers (e.g., Name, Department, Phone, Email).\n"
            "- Works with comma/semicolon/tab separated files.\n"
        )
        messagebox.showinfo("Help", message)

    def _show_context_menu(self, event: tk.Event) -> None:
        row_id = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)
        if row_id:
            self.tree.selection_set(row_id)
            self.tree.focus(row_id)
        self._context_item = row_id or None
        self._context_column = col_id or None

        menu = tk.Menu(self, tearoff=False)
        menu.add_command(label="Copy Phone", command=self.copy_phone, state="normal" if self.phone_column else "disabled")
        menu.add_command(label="Copy Row", command=self.copy_row, state="normal" if self._selected_record() else "disabled")

        cell_state = "normal" if (self._selected_record() and self._context_column) else "disabled"
        menu.add_command(label="Copy Cell", command=self._copy_context_cell, state=cell_state)

        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def _copy_context_cell(self) -> None:
        record = self._selected_record()
        if not record or not self._context_column:
            return
        try:
            col_index = int(self._context_column.replace("#", "")) - 1
        except Exception:
            return
        if col_index < 0 or col_index >= len(self.columns):
            return
        col = self.columns[col_index]
        value = record.get(col, "")
        self.clipboard_clear()
        self.clipboard_append(value)
        self._refresh_status(f"Copied {col}.")


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Teacher phone book (CSV viewer)")
    parser.add_argument("--csv", help="Path to teacher CSV file")
    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    set_windows_dpi_awareness()
    initial = detect_initial_csv_path(args.csv)
    app = PhoneBookApp(initial_csv=initial)
    app.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
