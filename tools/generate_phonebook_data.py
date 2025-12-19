import argparse
import csv
import json
from datetime import datetime, timezone
from io import StringIO
from pathlib import Path

import openpyxl


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate mobile/phonebook_data.js from PhoneBook.xlsx")
    parser.add_argument("--input", default="PhoneBook.xlsx", help="Input .xlsx file (default: PhoneBook.xlsx)")
    parser.add_argument(
        "--output",
        default=str(Path("mobile") / "phonebook_data.js"),
        help="Output JS file (default: mobile/phonebook_data.js)",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Sheet name to read (default: first sheet)",
    )
    return parser.parse_args(argv)


def read_xlsx_csv_lines(path: Path, sheet_name: str | None) -> list[str]:
    wb = openpyxl.load_workbook(path)
    name = sheet_name or wb.sheetnames[0]
    ws = wb[name]

    lines: list[str] = []
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row=row, column=1).value
        if value is None:
            continue
        text = str(value).strip()
        if not text:
            continue
        lines.append(text)
    if not lines:
        raise ValueError("No rows found in the first column.")
    return lines


def parse_csv_lines(lines: list[str]) -> tuple[list[str], list[dict[str, str]]]:
    reader = csv.reader(StringIO("\n".join(lines)))
    rows = list(reader)
    if not rows:
        raise ValueError("No CSV rows found.")
    headers = [h.strip() for h in rows[0]]
    if not any(headers):
        raise ValueError("Header row is empty.")

    records: list[dict[str, str]] = []
    for row in rows[1:]:
        item: dict[str, str] = {}
        for idx, header in enumerate(headers):
            value = row[idx] if idx < len(row) else ""
            item[header] = str(value).strip()
        if any(v for v in item.values()):
            records.append(item)
    return headers, records


def generate_js(data: list[dict[str, str]], source_name: str) -> str:
    generated_at = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    payload = json.dumps(data, ensure_ascii=False, indent=2)

    meta = {
        "generatedAt": generated_at,
        "source": source_name,
        "count": len(data),
    }
    meta_payload = json.dumps(meta, ensure_ascii=False, indent=2)

    return (
        "// Auto-generated file. You can edit it manually, but it may be overwritten by the generator.\n"
        f"// Source: {source_name}\n"
        f"// Generated: {generated_at}\n\n"
        f"window.PHONEBOOK_DATA = {payload};\n\n"
        f"window.PHONEBOOK_META = {meta_payload};\n"
    )


def main() -> int:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)

    if not input_path.exists():
        raise SystemExit(f"Input file not found: {input_path}")

    lines = read_xlsx_csv_lines(input_path, args.sheet)
    _headers, records = parse_csv_lines(lines)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(generate_js(records, input_path.name), encoding="utf-8")
    print(f"Wrote {len(records)} records to {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

